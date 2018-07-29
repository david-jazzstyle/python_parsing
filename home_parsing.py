import csv
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook


if __name__ == "__main__":
	print("Hello World")

	req_skin = requests.get("http://www.shilladfs.com/estore/kr/ko/Skin-Care/c/1?actr=1#url")
	html_skin = req_skin.text
	soup_skin = BeautifulSoup(html_skin, 'html.parser')

	req_makeup = requests.get("http://www.shilladfs.com/estore/kr/ko/Makeup/c/2?actr=1")
	html_makeup = req_makeup.text
	soup_makeup = BeautifulSoup(html_makeup, 'html.parser')


	# branddata = soup_skin.find_all("div", {"class" : "brand"})

	

	# my_brands = soup_skin.select(
	# 	# 'container > div.sub_category > div.product_list_wrap.facet_module > div.facet-product-list > ul:nth-child(4) > li:nth-child(1) > div > div.product_off > div.pr_info'
	# 	# 'container > div.sub_category > div.product_list_wrap.facet_module > div.facet-product-list > ul:nth-child(2) > li:nth-child(1) > div > div.product_off > div.pr_info > div.price > span.sale'
	# 	# 'div[class=brand]'

	# 	'[@id="container"]>div[2]>div[3]>div[3]>ul[4]>li[1]>div>div[1]>div[2]'
	# 	)
	# my_brands = soup_skin.find(id='container').find_all('div').find_all('ul').find_all('li').find('div')
	# //*[@id="container"]/div[2]/div[3]/div[3]/ul[4]/li[1]/div/div[1]/div[2]/div[1]
	# 1 2 2 3 0 div 0 1 0 

	# //*[@id="container"]/div[2]/div[3]/div[3]/ul[1]/li[1]/div/div[1]/div[2]/div[4]/span[1]

	# print(soup_skin.find(id='container').find_all('div')[1].find_all('div')[2].find_all('div')[2].find_all('ul')[0].find_all('li')[1])

	# //*[@id="container"]/div[2]/div[3]/div[3]/ul[1]/li[1]/div/div[1]/div[2]/div[4]/span[1]	
	# print(soup_skin.find(id='container').find_all('div')[1].find_all('div')[2].find_all('div')[2].find_all('ul')[0].find_all('li')[1])
	# print(soup_skin.find(id='container').find_all('div')[1].find_all('div')[2].find_all('div')[2].find_all('ul')[0].find_all('li')[0].find_all('div', {"class" : "pr_info"}))




	# for title in my_titles:
	# 	print(title.text)
	#//*[@id="container"]/div[2]/div[3]/div[3]/ul[1]/li[1]/div/div[1]/div[2]/div[4]/span[1]
	# print(soup_skin.find_all("div", {"class" : "facet-product-list"}).find_all("div", {"class" : "pr_info"}))
	# my_sales = soup_skin.select(
	# 	"ul li span.sale"
	# 	)

	my_sales = soup_skin.select(
		"ul.product_list_wr"
		)
	
	for sale in my_sales:
		print(sale)

	# my_names = soup_skin.select(
	# # 	# '#best_selling1 > div.best_box > ul > li.list_01 > div > div.product_off > div.pr_info > div.price > span.sale'
	# 	'div[class=name]'
	# 	)

	# my_sales = soup_skin.select(
	# # 	# '#best_selling1 > div.best_box > ul > li.list_01 > div > div.product_off > div.pr_info > div.price > span.sale'
	# 	'span[class=sale]'
	# 	)


	# wb = Workbook()
	# ws = wb.active

	# n = 1
	# for brand in my_brands:
	# 	# ws.cell(row=n, column=1, value=brand.text)
	# 	n = n+1
	# 	print(brand)

	# # n = 1
	# # for name in my_names:
	# # 	ws.cell(row=n, column=2, value=name.text)
	# # 	n = n+1

	# # n = 1
	# # for sale in my_sales:
	# # 	ws.cell(row=n, column=3, value=sale.text)
	# # 	n = n+1



	# wb.save('test.xlsx')
