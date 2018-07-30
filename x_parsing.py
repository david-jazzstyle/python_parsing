import openpyxl
import requests
from bs4 import BeautifulSoup
 
req_skin = requests.get("http://www.shilladfs.com/estore/kr/ko/Skin-Care/c/1?actr=1")
req_makeup = requests.get("http://www.shilladfs.com/estore/kr/ko/Makeup/c/2?actr=1")
req_bhf = requests.get("http://www.shilladfs.com/estore/kr/ko/Body-Hair-Fragrance/c/3?actr=1")
 
html_skin = req_skin.text
soup_skin = BeautifulSoup(html_skin, 'html.parser')
 
html_makeup = req_makeup.text
soup_makeup = BeautifulSoup(html_makeup, 'html.parser')
 
html_bhf = req_bhf.text
soup_bhf = BeautifulSoup(html_bhf, 'html.parser')
 
def fill_out(type_category, n):
          
           # ws = wb.active
           # "ws" + type_category
           ws = wb.create_sheet(index=n, title=type_category)
 
           if(type_category=="skin"):
                     soup_category = soup_skin
           elif(type_category=="makeup"):
                     soup_category = soup_makeup
           elif(type_category=="bhf"):
                     soup_category = soup_bhf
 
           my_brands = soup_category.select(
                     'div[class=brand]'
                     )
 
           my_names = soup_category.select(
                     'div[class=name]'
                     )
 
           my_sales = soup_category.select( 
                     'span[class=sale]'
                     )
 
           # [@id="container"]/div[2]/div[3]/div[3]/ul[1]/li[1]/div/div[1]/div[2]/div[4]/span[1]
           # my_sales = soup_category.find(id='container').find_all('div')[1].find_all('div')[2].find_all('div')[2].find_all('ul')[0].find_all('li')[0].find_all('div', {"class" : "pr_info"})
 
           n_row=1
           for brand in my_brands: 
                     ws.cell(row=n_row, column=1, value=brand.text)
                     n_row=n_row+1
 
           n_row=1
           for name in my_names:   
                     ws.cell(row=n_row, column=2, value=name.text)
                     n_row=n_row+1
 
           n_row=1
           for sale in my_sales:       
                     ws.cell(row=n_row, column=3, value=sale.text)
                     n_row=n_row+1
 
          
 
 
if __name__ == "__main__":
           print("Hello World")
 
           wb = openpyxl.Workbook()
           goods_type = ['skin', 'makeup', 'bhf']
           len_goods_type = len(goods_type)
           n = len(goods_type) -1
 
           while(n >= 0):
                     fill_out(goods_type[n], n)
                     n = n-1
          
           wb.remove(wb['Sheet'])
           wb.save(filename='test1.xlsx')